# -*- coding: utf-8 -*-
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

pts = json.load(open('points.json'))
prods = json.load(open('products.json'))['products']
term = json.load(open('terminal_products.json'))

biz = [f['properties'] for f in pts['features']] + \
      [(u['properties'] if 'properties' in u else u) for u in pts.get('unlocated', [])]
biz.sort(key=lambda r: r['seq'])

def norm(s): return re.sub(r'\s+', '', s or '').replace('　', '')
alias2maker = {}
for p in prods:
    comp = p['company']
    maker = ('泰山' if '泰山' in comp else '福壽' if '福壽' in comp else '福懋' if '福懋' in comp else comp)
    for k in [p['name']] + p.get('aliases', []):
        alias2maker[norm(k)] = maker

def maker_by_batch(b):
    b = (b or '').strip()
    if b.startswith('315-1150404'): return '中聯油脂(源頭)'
    if b[:2] in ('C1','C2') or b.startswith('BL'): return '福壽'
    if re.fullmatch(r'2026\d{8}', b): return '福懋'
    if re.fullmatch(r'2027\d{8,}', b): return '福懋'
    if re.fullmatch(r'2027\d{6}', b): return '泰山'
    return ''

def maker_for(prod, batch):
    n = norm(prod)
    if n in alias2maker: return alias2maker[n]
    if '益康' in (prod or ''): return '福懋'
    if '泰山' in (prod or ''): return '泰山'
    if '健味香油' in (prod or ''): return '福壽'
    if (batch or '').startswith('315-1150404'): return '中聯油脂(源頭)'
    if prod in ('原油', '大豆沙拉油'): return '中聯油脂(源頭)'
    if prod == '一級黃豆油':
        return '中聯油脂(源頭)' if (batch or '').startswith('315') else '福懋'
    return maker_by_batch(batch)

notes = {}
for s in (98, 331): notes[s] = '※官方註記：同序號96「建樹企業有限公司」（重複）'
for s in (104, 243, 298, 345): notes[s] = '※官方註記：經衛生局回復應刪除（第二批名單已更新，未販售該批問題油品給4家業者）'
for s in (240, 266, 267, 301, 322): notes[s] = '※官方註記：飼料用（非供食品烹製）'
notes[274] = '※官方註記：作「環氧大豆油」增塑劑/安定劑用途，非供食品烹製'
notes[360] = '※官方註記：單據誤植「誠一」，實際應為「和香行」，地址品項不變'

HDR = PatternFill('solid', fgColor='1F4E78'); HDRF = Font(color='FFFFFF', bold=True, size=11)
thin = Side(style='thin', color='BFBFBF'); BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical='top'); TOP = Alignment(vertical='top')
mk_fill = {'泰山': 'FCE4D6', '福壽': 'E2EFDA', '福懋': 'DDEBF7', '中聯油脂(源頭)': 'FFF2CC'}

def style_header(ws, ncol, row=1):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.border = BORD
        cell.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = f"A{row}:{get_column_letter(ncol)}{row}"

wb = Workbook()
ws = wb.active; ws.title = '事件摘要'; ws.sheet_view.showGridLines = False
t = ws.cell(1,1,'中聯油脂「問題大豆沙拉油」事件——商家與產品比對總表'); t.font = Font(size=16, bold=True, color='1F4E78'); ws.merge_cells('A1:C1')
s = ws.cell(2,1,'資料以食品藥物管理署（TFDA）「中聯油脂案專區」官方公告為準'); s.font = Font(size=10, italic=True, color='808080'); ws.merge_cells('A2:C2')
facts = [
 ('項目','內容','官方/來源'),
 ('事件','中聯油脂供應之部分大豆沙拉油檢出一級致癌物苯（a）駢芘（BaP）不符規定','食藥署'),
 ('污染程度','自檢 BaP 8.1 μg/kg，為限量標準 2.0 μg/kg 的逾 4 倍','食藥署/中央社'),
 ('問題源頭批號','大豆沙拉油／原油／一級黃豆油 批號 315-1150404（有效日期 115.9.30）','食藥署附件'),
 ('流向數量','約 1,300 公噸流向第一層三大廠：福懋約588噸、福壽約421噸、泰山約291噸','食藥署/中央社'),
 ('第一層業者','泰山企業、福壽實業、福懋油脂（三家均分中聯33.33%股權之代工平台股東）','食藥署'),
 ('受影響上游產品','18 項受影響產品、30 個批號（本表「上游受影響產品」分頁列品項/批號）','食藥署 7/6'),
 ('下游業者','360 家下游業者（截至 115/7/8 09:00 官方清單）','食藥署 7/7'),
 ('終端下架產品','預防性下架產品清單：截至 115/7/8 共 401 項（7/7 首批 232 項）','食藥署 7/8'),
 ('通報時序爭議','中聯 6/11 接獲檢驗異常，遲至 6/30 才通報，涉隱匿、延遲通報、虛報產量','食藥署/檢方'),
 ('裁罰','依食安法合計裁罰 1 億 6,520 萬元，創國內食安案史上最高','食藥署'),
 ('主管處置','要求下架回收與預防性下架；4大精煉廠採逐船/逐批檢驗 BaP','食藥署 7/6'),
]
for i,row in enumerate(facts):
    for j,val in enumerate(row):
        c = ws.cell(4+i, j+1, val); c.border = BORD; c.alignment = WRAP
        if i==0: c.fill=HDR; c.font=HDRF
ws.column_dimensions['A'].width=16; ws.column_dimensions['B'].width=78; ws.column_dimensions['C'].width=18
ws.freeze_panes='A5'

TS_BATCH_FIX={'20270408':'2027040801','20270409':'2027040901','20270725d':'2027072506','20270413E':'2027041301','20271013':'2027101301','20270725f':'2027072504','20271009':'2027100901'}
ws = wb.create_sheet('上游受影響產品(17項)')
hdr=['製造商','產品名稱','批號','有效日期','別名(通報名稱)']; ws.append(hdr)
for p in prods:
    comp=p['company']; maker=('泰山' if '泰山' in comp else '福壽' if '福壽' in comp else '福懋' if '福懋' in comp else comp)
    for b in p['batches']:
        code=TS_BATCH_FIX.get(b['code'], b['code'])
        ws.append([maker, p['name'], code, b['expiry'], '、'.join(p.get('aliases',[]))])
ws.append(['中聯油脂(源頭)','大豆沙拉油／原油／一級黃豆油','315-1150404','115.9.30 (2026.09.30)','問題源頭批號'])
style_header(ws,len(hdr))
for r in range(2,ws.max_row+1):
    mk=ws.cell(r,1).value
    for c in range(1,len(hdr)+1):
        cell=ws.cell(r,c); cell.border=BORD; cell.alignment=TOP
    if mk in mk_fill:
        for c in range(1,len(hdr)+1): ws.cell(r,c).fill=PatternFill('solid',fgColor=mk_fill[mk])
for col,w in zip('ABCDE',[16,34,18,20,30]): ws.column_dimensions[col].width=w

ws = wb.create_sheet('下游業者-360家')
hdr=['序號','縣市','業者','購買品項','批號','有效日期','來源製造商','官方備註']; ws.append(hdr)
for b in biz:
    pl=b.get('products',[]); bs=b.get('batches',[]); ex=b.get('expiries',[])
    mks=[maker_for(pl[i], bs[i] if i<len(bs) else '') for i in range(len(pl))]
    mku='、'.join(dict.fromkeys([m for m in mks if m]))
    ws.append([b['seq'],'/'.join(b.get('counties',[])),b['name'],' ｜ '.join(pl),' ｜ '.join(bs),' ｜ '.join(ex),mku,notes.get(b['seq'],'')])
style_header(ws,len(hdr))
for r in range(2,ws.max_row+1):
    for c in range(1,len(hdr)+1):
        cell=ws.cell(r,c); cell.border=BORD; cell.alignment=WRAP
    if ws.cell(r,8).value: ws.cell(r,8).fill=PatternFill('solid',fgColor='FFF2CC')
for col,w in zip('ABCDEFGH',[6,10,30,34,26,20,16,40]): ws.column_dimensions[col].width=w

ws = wb.create_sheet('下游業者x品項(展開)')
hdr=['序號','縣市','業者','購買品項','批號','有效日期','來源製造商','官方備註']; ws.append(hdr)
for b in biz:
    pl=b.get('products',[]); bs=b.get('batches',[]); ex=b.get('expiries',[])
    for i in range(len(pl)):
        bt=bs[i] if i<len(bs) else ''; e=ex[i] if i<len(ex) else ''
        ws.append([b['seq'],'/'.join(b.get('counties',[])),b['name'],pl[i],bt,e,maker_for(pl[i],bt),notes.get(b['seq'],'')])
style_header(ws,len(hdr))
for r in range(2,ws.max_row+1):
    for c in range(1,len(hdr)+1):
        cell=ws.cell(r,c); cell.border=BORD; cell.alignment=TOP
    mk=ws.cell(r,7).value
    if mk in mk_fill: ws.cell(r,7).fill=PatternFill('solid',fgColor=mk_fill[mk])
    if ws.cell(r,8).value: ws.cell(r,8).fill=PatternFill('solid',fgColor='FFF2CC')
for col,w in zip('ABCDEFGH',[6,10,30,30,18,14,16,40]): ws.column_dimensions[col].width=w

ws = wb.create_sheet('終端下架產品(401項)')
hdr=['業者序號','縣市','業者','產品序號','終端產品名稱','有效日期/狀態']; ws.append(hdr)
for v in term['vendors']:
    for p in v['products']:
        ws.append([v['seq'],v.get('county',''),v['name'],p.get('seq',''),p['name'],p.get('expiry','')])
style_header(ws,len(hdr))
for r in range(2,ws.max_row+1):
    for c in range(1,len(hdr)+1):
        cell=ws.cell(r,c); cell.border=BORD; cell.alignment=TOP
for col,w in zip('ABCDEF',[10,10,32,10,46,18]): ws.column_dimensions[col].width=w

ws = wb.create_sheet('資料來源與說明'); ws.sheet_view.showGridLines=False
lines=[
 ('資料來源與說明',True),('',False),
 ('1. 主要一手來源：衛福部食品藥物管理署「中聯油脂案專區」',False),
 ('   https://www.fda.gov.tw/tc/site.aspx?sid=13702',False),
 ('   ├ 下游業者360家清單（截至7月8日）PDF',False),
 ('   └ 受影響油品資訊 PDF（食藥署 7/6 公告，18項/30批號）',False),
 ('2. 食藥署新聞稿：五項強化措施‧公布更新360家業者及產品資訊（2026-07-06）',False),
 ('   https://www.fda.gov.tw/TC/newsContent.aspx?cid=4&id=t634424',False),
 ('3. 食藥署附件：中聯油脂等4家業者通報產品資訊',False),
 ('   https://www.fda.gov.tw/tc/includes/GetFile.ashx?id=t408957',False),
 ('4. 結構化整理/地理定位/終端產品彙整：tainan.olc.tw「問題油品流向地圖」',False),
 ('   （資料源標註為食藥署中聯油脂案專區）https://tainan.olc.tw/p/oil_products/',False),
 ('5. 事件數據佐證：中央社 CNA（2026/7/6）',False),
 ('',False),('欄位說明：',True),
 ('‧「來源製造商」依產品/批號對應第一層製造商（泰山/福壽/福懋）或源頭中聯油脂。',False),
 ('‧「益康」品牌為福懋油脂；健味香油/沙拉油18L/塑桶3L為福壽；金酥耐炸油等為泰山。',False),
 ('‧「官方備註」為食藥署360家清單PDF之官方註腳。',False),
 ('',False),('重要提醒：',True),
 ('‧本表為115/7/8官方公告快照，名單持續更新，請以食藥署專區最新版為準。',False),
 ('‧列入名單代表進貨端曾涉問題油品流向，不必然代表最終販售產品仍在架或有害。',False),
 ('‧序號104/243/298/345應刪除；98/331與96重複；240/266/267/301/322飼料用；274工業用。',False),
 ('‧完整性註記：360家業者名冊、縣市、序號均與官方PDF逐筆核對一致（座標解析）。',False),
 ('‧少數業者進貨含多種品項；本表記錄其主要品項與該品項全部批號，個別次要品項之逐項',False),
 ('  明細以食藥署「下游業者360家清單」PDF為準（見上方連結）。',False),
 ('‧終端下架產品清單狀態：'+term['updated'],False),
]
for i,(txt,bold) in enumerate(lines,1):
    c=ws.cell(i,1,txt); c.font=Font(bold=bold,size=13 if (bold and i==1) else 11,color='1F4E78' if bold else '000000')
ws.column_dimensions['A'].width=100

out='中聯油脂問題油品_商家與產品比對_官方公告.xlsx'; wb.save(out)
from collections import Counter
mk=Counter(); rows=0
for b in biz:
    for i,pr in enumerate(b.get('products',[])):
        bt=b['batches'][i] if i<len(b.get('batches',[])) else ''
        mk[maker_for(pr,bt)]+=1; rows+=1
print('SAVED',out)
print('businesses',len(biz),'expanded_rows',rows)
print('maker_dist',dict(mk))
print('terminal_vendors',len(term['vendors']),'terminal_products',sum(len(v['products']) for v in term['vendors']))
print('unmatched_maker',mk.get('',0))
